import streamlit as st
import pandas as pd
import re
from io import BytesIO
from fpdf import FPDF
# (Import SchedulerLogic ของคุณตามปกติ)
from scheduler_logic import SchedulerCSP 
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# ... (ส่วน Config และ CSS คงเดิม) ...

# ==========================================
# 3. Helper Functions
# ==========================================
def clean_teacher_name(name):
    """ตัดคำนำหน้าชื่อเพื่อให้สั้นลงและพอดีช่อง"""
    if pd.isna(name): return ""
    s = str(name).strip()
    # รายการคำนำหน้าที่จะตัดออก
    prefixes = [
        'ว่าที่ร้อยตรี', 'ว่าที่ ร.ต.', 'ว่าที่ร.ต.', 'ดร.', 'ผศ.', 'รศ.', 'ศ.',
        'นางสาว', 'นาย', 'นาง', 'มิส', 'มาสเตอร์', 
        'Mr.', 'Mrs.', 'Miss.', 'Ms.', 'Master', 'Teacher'
    ]
    for p in prefixes:
        s = s.replace(p, '')
    return s.strip()

# ==========================================
# 4. New PDF Engine (Pixel-Perfect Layout)
# ==========================================
class PDF(FPDF):
    def footer(self):
        self.set_y(-12)
        self.set_font('THSarabunNew', '', 10)
        self.set_text_color(100, 100, 100)
        self.cell(0, 10, f'หน้า {self.page_no()}', 0, 0, 'R')

def gen_pdf(df, entities, vkey, t_map):
    # ตั้งค่าหน้ากระดาษ A4 แนวนอน (297 x 210 mm)
    pdf = PDF('L', 'mm', 'A4')
    pdf.set_auto_page_break(False) # ปิด Auto Break เพื่อคุม Layout เอง
    
    # *** สำคัญ: ต้องมีไฟล์ฟอนต์ THSarabunNew.ttf ในโฟลเดอร์เดียวกัน ***
    try:
        pdf.add_font('THSarabunNew', '', 'THSarabunNew.ttf', uni=True)
        pdf.add_font('THSarabunNew', 'B', 'THSarabunNew Bold.ttf', uni=True) # ถ้ามีตัวหนา
    except:
        pdf.add_font('Arial', '', 10) # Fallback
    
    cfg = VIEWS[vkey]
    
    # --- Configuration ขนาดช่อง (รวมกันต้องไม่เกิน 277mm) ---
    MARGIN_LEFT = 10
    MARGIN_TOP = 15
    W_DAY = 25        # ช่องวัน
    W_SLOT = 28       # ช่องคาบเรียน (28mm * 8 คาบ = 224mm)
    W_LUNCH = 20      # ช่องพักเที่ยง
    H_HEADER = 16     # ความสูงหัวตาราง
    H_ROW = 22        # ความสูงแถวตาราง
    
    entities_list = [entities] if isinstance(entities, str) else entities
    
    for ent in entities_list:
        sub = df[df[cfg['id']] == ent]
        if sub.empty: continue
        
        pdf.add_page()
        
        # --- 1. Title ---
        title_name = t_map.get(ent, ent) if vkey=='Teacher' else ent
        if vkey == 'Teacher': title_name = clean_teacher_name(title_name) # Clean ชื่อหัวกระดาษด้วย
        
        pdf.set_font('THSarabunNew', 'B', 20) # ตัวหนา
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, f"ตารางสอน: {title_name}", 0, 1, 'C')
        pdf.ln(5)
        
        # --- 2. Draw Header (วาดทีละกล่อง) ---
        start_y = pdf.get_y()
        curr_x = MARGIN_LEFT
        
        pdf.set_font('THSarabunNew', 'B', 12)
        pdf.set_fill_color(27, 94, 32) # สีเขียวเข้ม
        pdf.set_text_color(255, 255, 255) # ขาว
        
        # 2.1 หัวข้อ "วัน/เวลา"
        pdf.set_xy(curr_x, start_y)
        pdf.cell(W_DAY, H_HEADER, "วัน / เวลา", 1, 0, 'C', 1)
        curr_x += W_DAY
        
        # 2.2 วนลูปคาบเรียน
        # สมมติ Periods = [1, 2, 3, 4, 'Lunch', 5, 6, 7, 8]
        # ต้องแน่ใจว่า PERIODS ถูก define ไว้ข้างนอก หรือ pass เข้ามา
        periods_layout = [1, 2, 3, 4, 'Lunch', 5, 6, 7, 8] 
        
        for p in periods_layout:
            w = W_LUNCH if p == 'Lunch' else W_SLOT
            
            # วาดพื้นหลังกล่อง
            pdf.set_xy(curr_x, start_y)
            pdf.cell(w, H_HEADER, "", 1, 0, 'C', 1)
            
            # เตรียมข้อความ
            if p == 'Lunch':
                line1 = "12:30 - 13:30"
                line2 = "พักกลางวัน"
            else:
                line1 = TIME_MAP.get(p, "") # ดึงเวลาจาก Config
                line2 = f"คาบที่ {p}"
            
            # เขียนเวลา (ตัวเหลือง, บรรทัดบน)
            pdf.set_text_color(255, 241, 118) # เหลืองอ่อน
            pdf.set_xy(curr_x, start_y + 2)
            pdf.cell(w, 5, line1, 0, 2, 'C')
            
            # เขียนคาบ (ตัวขาว, บรรทัดล่าง)
            pdf.set_text_color(255, 255, 255)
            pdf.set_xy(curr_x, start_y + 8) # ขยับลงมา
            pdf.cell(w, 5, line2, 0, 0, 'C')
            
            curr_x += w
            
        pdf.set_text_color(0, 0, 0) # Reset สีดำ
        current_y = start_y + H_HEADER
        
        # --- 3. Draw Grid Rows ---
        days_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
        days_th = {'Mon':'จันทร์', 'Tue':'อังคาร', 'Wed':'พุธ', 'Thu':'พฤหัสบดี', 'Fri':'ศุกร์'}
        
        for d in days_order:
            curr_x = MARGIN_LEFT
            
            # 3.1 ช่องวัน (ซ้ายสุด)
            pdf.set_font('THSarabunNew', 'B', 14)
            pdf.set_fill_color(232, 245, 233) # เขียวอ่อนมาก
            pdf.set_xy(curr_x, current_y)
            pdf.cell(W_DAY, H_ROW, days_th[d], 1, 0, 'C', 1)
            curr_x += W_DAY
            
            # 3.2 ช่องข้อมูล
            for p in periods_layout:
                w = W_LUNCH if p == 'Lunch' else W_SLOT
                pdf.set_xy(curr_x, current_y)
                
                if p == 'Lunch':
                    pdf.set_fill_color(224, 224, 224) # เทา
                    pdf.cell(w, H_ROW, "พัก", 1, 0, 'C', 1)
                else:
                    # หาข้อมูลใน DataFrame
                    r = sub[(sub['Day'] == d) & (sub['Period'] == p)]
                    
                    pdf.set_fill_color(255, 255, 255) # ขาว
                    pdf.rect(curr_x, current_y, w, H_ROW) # วาดกรอบ
                    
                    if not r.empty:
                        row_data = r.iloc[0]
                        # ดึงข้อมูลตาม Config มุมมอง
                        val1 = str(row_data[cfg['cols'][0]]) # เช่น ห้อง
                        val2 = str(row_data[cfg['cols'][1]]) # เช่น รหัสวิชา
                        val3 = str(row_data[cfg['cols'][2]]) # เช่น ชื่อครู
                        
                        # Clean ชื่อครูถ้าจำเป็น
                        if 'Teacher' in cfg['cols'][2]:
                            val3 = clean_teacher_name(val3)
                        
                        # ตัดคำถ้ายาวเกิน (Truncate)
                        val1 = val1[:15]
                        val2 = val2[:15]
                        val3 = val3[:18]
                        
                        # เขียน 3 บรรทัด
                        pdf.set_font('THSarabunNew', '', 10)
                        
                        # บรรทัด 1 (ห้อง/บนสุด)
                        pdf.set_xy(curr_x, current_y + 2)
                        pdf.cell(w, 5, val1, 0, 0, 'C')
                        
                        # บรรทัด 2 (รหัสวิชา/กลาง)
                        pdf.set_xy(curr_x, current_y + 7.5)
                        pdf.set_font('THSarabunNew', 'B', 10) # ตัวหนา
                        pdf.cell(w, 5, val2, 0, 0, 'C')
                        pdf.set_font('THSarabunNew', '', 10)
                        
                        # บรรทัด 3 (ครู/ล่างสุด)
                        pdf.set_xy(curr_x, current_y + 13)
                        pdf.cell(w, 5, val3, 0, 0, 'C')
                
                curr_x += w
            current_y += H_ROW # จบแถว ขยับ Y ลง
            
        # --- 4. Legend (Compact) ---
        current_y += 5
        pdf.set_xy(MARGIN_LEFT, current_y)
        pdf.set_font('THSarabunNew', 'B', 12)
        pdf.cell(0, 8, "รายละเอียดรายวิชา:", 0, 1, 'L')
        current_y += 8
        
        # Legend Header
        pdf.set_fill_color(200, 230, 201) # เขียวอ่อน
        wds = [25, 90, 40, 50] # ความกว้างคอลัมน์ Legend
        pdf.set_xy(MARGIN_LEFT, current_y)
        headers = cfg['leg'] # ['รหัส', 'ชื่อวิชา', 'ห้อง', 'ครู']
        for i, h in enumerate(headers):
            pdf.cell(wds[i], 7, h, 1, 0, 'C', 1)
        current_y += 7
        
        # Legend Rows
        pdf.set_font('THSarabunNew', '', 11)
        leg_df = sub[cfg['leg_c']].drop_duplicates()
        
        for _, row in leg_df.iterrows():
            # เช็คหน้าหมด
            if current_y > 180:
                pdf.add_page(); current_y = 20
            
            pdf.set_xy(MARGIN_LEFT, current_y)
            # col 0: ID
            pdf.cell(wds[0], 7, str(row[0]), 1, 0, 'C')
            # col 1: Name (Left align)
            pdf.cell(wds[1], 7, str(row[1])[:60], 1, 0, 'L')
            # col 2: Room
            pdf.cell(wds[2], 7, str(row[2]), 1, 0, 'C')
            # col 3: Teacher/Group (Clean Name)
            t_val = str(row[3])
            if 'Teacher' in cfg['leg_c'][3]: t_val = clean_teacher_name(t_val)
            pdf.cell(wds[3], 7, t_val[:30], 1, 0, 'L')
            
            current_y += 7
            
    return pdf.output(dest='S').encode('latin-1')

def gen_excel(df, t_map):
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Raw')
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for k, cfg in VIEWS.items():
            col = f'Disp_{k}'; df[col] = df[cfg['cols'][0]] + "\n" + df[cfg['cols'][1]] + "\n" + df[cfg['cols'][2]]
            for ent in sorted(df[cfg['id']].unique()):
                sub = df[df[cfg['id']] == ent]
                if sub.empty: continue
                piv = sub.pivot_table(index='Day', columns='Period', values=col, aggfunc='first').reindex(DAYS_EN).reindex(columns=[1,2,3,4,5,6,7,8])
                try: piv.insert(4, 'Lunch', 'พักกลางวัน') 
                except: pass
                piv.index = piv.index.map(DAY_MAP)
                piv.columns = [TIME_MAP.get(c, str(c)).replace("\n", " ") if c!='Lunch' else "12:30-13:30" for c in piv.columns]
                
                sh_name = f"{cfg['pfx']}{str(ent)[:20]}".replace(":","").replace("/","-")
                piv.fillna('').to_excel(writer, sheet_name=sh_name)
                
                ws = writer.sheets[sh_name]
                ws.column_dimensions['A'].width = 15
                for c in range(2, 12): ws.column_dimensions[chr(64+c)].width = 25; ws.cell(row=1, column=c).alignment = align
                for row in ws.iter_rows():
                    for cell in row: cell.alignment = align; cell.border = thin
    return out.getvalue()

# ==========================================
# 5. MAIN UI
# ==========================================
st.sidebar.header("1. นำเข้าข้อมูล")
up = st.sidebar.file_uploader("Upload CSV/Excel", accept_multiple_files=True)

if up:
    data, logs = load_data(up)
    if logs:
        with st.sidebar.expander("🛠️ บันทึกการตรวจสอบ (Validation)", expanded=True):
            for l in logs:
                if "ลบ" in l or "ตัด" in l: st.warning(l, icon="🧹")
                elif "Error" in l: st.error(l)
                else: st.info(l)
            
    if len(data) == 4:
        st.sidebar.success("✅ ข้อมูลพร้อม")
        t_map = dict(zip(data['Teachers']['TeacherID'], data['Teachers']['CleanName']))
        
        if st.sidebar.button("🚀 สร้างตาราง"):
            with st.spinner("กำลังจัดตารางสอน (AI)..."):
                res, una = SchedulerCSP(data['Teachers'], data['Subjects'], data['Rooms'], data['Groups']).generate_schedule(45)
                if [i for l in res.values() for i in l]:
                    df = pd.DataFrame([i for l in res.values() for i in l])
                    df['Teacher_Name'] = df['Teacher_ID'].map(t_map).fillna(df['Teacher_ID'])
                    st.session_state.update(res=df, una=una, t_map=t_map)
                    if not una: st.success("🎉 สำเร็จ 100%")
                    else: st.warning(f"⚠️ ตกหล่น {len(una)} รายการ")
                else: st.error("❌ ล้มเหลว")
    else: st.sidebar.warning(f"ไฟล์ไม่ครบ: {set(['Groups','Rooms','Teachers','Subjects']) - data.keys()}")

if 'res' in st.session_state:
    df, t_map = st.session_state.res, st.session_state.t_map
    if st.session_state.una: st.expander("รายการตกหล่น").write(st.session_state.una)
    st.divider()
    
    # Preview
    c1, c2 = st.columns([1, 4])
    vkey = c1.radio("มุมมอง:", list(VIEWS.keys()), format_func=lambda x: VIEWS[x]['lbl'])
    cfg = VIEWS[vkey]
    ents = sorted(df[cfg['id']].unique())
    sel = c1.selectbox("เลือกรายการ:", ents, format_func=(lambda x: t_map.get(x,x)) if vkey=='Teacher' else (lambda x: x))
    
    if sel:
        sub = df[df[cfg['id']] == sel].copy()
        sub['Disp'] = sub[cfg['cols'][0]] + "<br>" + sub[cfg['cols'][1]] + "<br>" + sub[cfg['cols'][2]]
        piv = sub.pivot_table(index='Day', columns='Period', values='Disp', aggfunc='first').reindex(DAYS_EN).fillna("-")
        
        # HTML Table
        h = "<table class='custom-table'><thead><tr><th>วันที่</th>"
        for p in PERIODS:
            if p == 'Lunch': t_str, lbl = "12:30 - 13:30", "พักกลางวัน"
            else: t_str, lbl = TIME_MAP.get(p, ""), f"คาบ {p}"
            h += f"<th><span class='time-txt'>{t_str}</span><span class='period-txt'>{lbl}</span></th>"
        h += "</tr></thead><tbody>"
        
        for d in DAYS_EN:
            h += f"<tr><td class='day-cell'>{DAY_MAP[d]}</td>"
            for p in PERIODS:
                v = "พัก" if p=='Lunch' else (piv.at[d,p] if p in piv.columns and pd.notna(piv.at[d,p]) else "-")
                bg = "background:#eee;" if p=='Lunch' else ""
                val = v if p!='Lunch' else "พัก"
                h += f"<td style='{bg}'>{val}</td>"
            h += "</tr>"
        h += "</tbody></table>"
        
        c2.markdown(f"### {t_map.get(sel,sel) if vkey=='Teacher' else sel}")
        c2.markdown(h, unsafe_allow_html=True)
        
        # Legend
        c2.markdown("#### ℹ️ รายละเอียดรายวิชา")
        ref_df = sub[cfg['leg_c']].drop_duplicates()
        ref_df.columns = cfg['leg']
        c2.table(ref_df)
        c2.download_button("📄 PDF หน้านี้", gen_pdf(df, sel, vkey, t_map), f"{sel}.pdf", "application/pdf")

    st.divider(); st.subheader("💾 ดาวน์โหลดทั้งหมด")
    
    # --- TABS: Download by Major ---
    tab1, tab2 = st.tabs(["📁 รวมเล่ม (All)", "📂 แยกสาขา (By Major)"])
    
    with tab1:
        cols = st.columns(4)
        cols[0].download_button("📥 Excel รวม", gen_excel(df, t_map), "Master.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        for i, (k, v) in enumerate(VIEWS.items()):
            if cols[i+1].button(f"📄 PDF {v['lbl'].split('(')[0]}"):
                with st.spinner("Generating..."):
                    st.session_state[f'p_{k}'] = gen_pdf(df, sorted(df[v['id']].unique()), k, t_map)
            if f'p_{k}' in st.session_state: cols[i+1].download_button("⬇️ โหลด", st.session_state[f'p_{k}'], f"{k}s.pdf")

    with tab2:
        st.info("💡 ดาวน์โหลด PDF แยกตามสาขาวิชา")
        cats = get_categories(df, 'Group_ID')
        cat_cols = st.columns(4)
        for i, (cat, items) in enumerate(cats.items()):
            with cat_cols[i % 4]:
                if st.button(f"📄 สาขา {cat}"):
                    st.session_state[f'pdf_{cat}'] = gen_pdf(df, sorted(items), 'Student', t_map)
                if f'pdf_{cat}' in st.session_state:
                    st.download_button(f"⬇️ {cat}.pdf", st.session_state[f'pdf_{cat}'], f"{cat}.pdf")

